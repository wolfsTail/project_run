from openpyxl.reader.excel import load_workbook
from rest_framework.decorators import api_view
from rest_framework.filters import SearchFilter
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.filters import OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework import viewsets, generics
from rest_framework import status

from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from django.http import Http404
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
from django.db.models import Count, Q, Sum

from haversine import haversine, Unit

from .models import AthleteInfo, Challenge, Position, Run, CollectibleItem
from .serializers import (
    RunSerializer, 
    UserSerializer, 
    AthleteInfoSerializer, 
    ChallengeSerializer,
    PositionSerializer,
    CollectibleItemSerializer,
)



TEN_RUNS_CHALLENGE = "Сделай 10 Забегов!"
FIFTY_KM_CHALLENGE = "Пробеги 50 километров!"

@api_view(['GET'])
def contacts_view(request):
    return Response(
        {
            'company_name': 'Рога и панцирь',
            'slogan': 'Бегаем или за кем-то или от кого-то',
            'contacts': 'NDA'
        }
    )


class PagePagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'size'
    max_page_size = 50

class OptionalPagePagination(PageNumberPagination):
    page_size = None
    page_size_query_param = 'size'
    max_page_size = 50

    def get_page_size(self, request):
        if self.page_size_query_param in request.query_params:
            try:
                return self._get_size_from_query(request)
            except ValueError:
                return None
        return None

    def _get_size_from_query(self, request):
        raw = request.query_params[self.page_size_query_param]
        size = int(raw)
        if size <= 0:
            raise ValueError("size must be positive")
        if self.max_page_size:
            size = min(size, self.max_page_size)
        return size


class StartRunApiView(APIView):
    def post(self, request, run_id):
        current = get_object_or_404(Run, id=run_id)

        if current.status != Run.STATUS_CHOICES[0][0]:
            return Response(
                {
                    "detail": "Run already started or finished!",
                    "id": current.id,
                    "status": current.status,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        current.status = Run.STATUS_CHOICES[1][0]
        current.save(update_fields=["status"])

        return Response(
            {
                "id": current.id,
                "status": current.status,
            },
            status=status.HTTP_200_OK,
        )


class StopRunApiView(APIView):
    def post(self, request, run_id):
        current = get_object_or_404(Run, id=run_id)

        if current.status != Run.STATUS_CHOICES[1][0]:
            return Response(
                {
                    "detail": "Run already started or finished!",
                    "id": current.id,
                    "status": current.status,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            current.status = Run.STATUS_CHOICES[2][0]
            current.save(update_fields=["status"])

            finished_count = Run.objects.filter(
                athlete=current.athlete, status="finished"
            ).count()

            if finished_count == 10 and not Challenge.objects.filter(
                athlete=current.athlete, full_name=TEN_RUNS_CHALLENGE
            ).exists():
                Challenge.objects.create(
                    athlete=current.athlete, full_name=TEN_RUNS_CHALLENGE
                )

            pts = list(
                Position.objects
                .filter(run=current)
                .order_by("created_at", "id")
                .values_list("latitude", "longitude")
            )

            distance_km = 0.0
            if len(pts) >= 2:
                prev = (float(pts[0][0]), float(pts[0][1]))
                for lat, lon in pts[1:]:
                    cur = (float(lat), float(lon))
                    distance_km += haversine(prev, cur, unit=Unit.KILOMETERS)
                    prev = cur

            current.distance = round(distance_km, 4)
            current.save(update_fields=["distance"])

            total_km = Run.objects.filter(
                athlete=current.athlete, status="finished"
            ).aggregate(s=Sum("distance"))["s"] or 0.0

            if total_km >= 50 and not Challenge.objects.filter(
                athlete=current.athlete, full_name=FIFTY_KM_CHALLENGE
            ).exists():
                Challenge.objects.create(athlete=current.athlete, full_name=FIFTY_KM_CHALLENGE)

        return Response(
            {
                "id": current.id,
                "status": current.status,
                "distance": current.distance,
            },
            status=status.HTTP_200_OK,
        )


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    pagination_class = OptionalPagePagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    pagination_class = OptionalPagePagination
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['=first_name', '=last_name']
    ordering_fields = ['date_joined']

    def get_queryset(self):
        qs = self.queryset.exclude(is_superuser=True)

        t = self.request.query_params.get('type')
        if t == 'coach':
            qs = qs.filter(is_staff=True)
        elif t == 'athlete':
            qs = qs.filter(is_staff=False)

        return qs


class AthleteInfoView(APIView):
    def get_user(self, user_id: int) -> User:
        return get_object_or_404(User, pk=user_id)

    def get(self, request, user_id: int):
        user = self.get_user(user_id)
        info, _ = AthleteInfo.objects.get_or_create(user=user)
        data = AthleteInfoSerializer(info).data
        return Response(data, status=status.HTTP_200_OK)

    def put(self, request, user_id: int):
        user = self.get_user(user_id)
        info, _ = AthleteInfo.objects.get_or_create(user=user)
        serializer = AthleteInfoSerializer(instance=info, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.select_related("athlete").order_by("-created_at")
    serializer_class = ChallengeSerializer
    pagination_class = OptionalPagePagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["athlete"]


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.select_related("run").order_by("id")
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["run"]
    http_method_names = ["get", "post", "delete", "head", "options"]

    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()
        run_id = self.request.query_params.get("run")
        if run_id:
            qs = qs.filter(run_id=run_id)
        return qs


class CollectibleItemListView(generics.ListAPIView):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadFileView(APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided under 'file'."},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(filename=file, read_only=True, data_only=True)
        except Exception as e:
            return Response({"detail": f"Failed to read workbook: {e}"},
                            status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            return Response([], status=status.HTTP_200_OK)

        header_map = {str(h).strip(): idx for idx, h in enumerate(header)}
        required = ["Name", "UID", "Value", "Latitude", "Longitude", "URL"]
        missing = [h for h in required if h not in header_map]
        if missing:
            invalid_rows = [list(r) for r in rows_iter if r and any(c is not None for c in r)]
            return Response(invalid_rows, status=status.HTTP_200_OK)

        invalid_rows = []
        valid_instances = []
        seen_uids = set()

        for row in rows_iter:
            if row is None or all(cell is None for cell in row):
                continue

            data = {
                "name": row[header_map["Name"]],
                "uid": row[header_map["UID"]],
                "value": row[header_map["Value"]],
                "latitude": row[header_map["Latitude"]],
                "longitude": row[header_map["Longitude"]],
                "picture": row[header_map["URL"]],
            }

            uid_val = str(data.get("uid") or "").strip()
            if uid_val in seen_uids:
                invalid_rows.append(list(row))
                continue
            if uid_val:
                seen_uids.add(uid_val)

            ser = CollectibleItemSerializer(data=data)
            if ser.is_valid():
                valid_instances.append(CollectibleItem(**ser.validated_data))
            else:
                invalid_rows.append(list(row))

        with transaction.atomic():
            for inst in valid_instances:
                try:
                    inst.save()
                except Exception:
                    invalid_rows.append([
                        inst.name, inst.uid, inst.value, inst.latitude, inst.longitude, inst.picture
                    ])

        return Response(invalid_rows, status=status.HTTP_200_OK)
